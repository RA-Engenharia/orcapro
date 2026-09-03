# -*- coding: utf-8 -*-
# Gera a proposta PC-2026-0902-01 (MY Engenharia) a partir do SINAPI-MG 06/2026 em app/data.
# Uso (na raiz do repo): python3 propostas/gerar-proposta-my-engenharia.py
# Depois: abrir o .html e "Imprimir / Salvar PDF", ou renderizar com Chromium headless.
import json, re, unicodedata, math, html, datetime
d = json.load(open('app/data/sinapi-MG-analitico.json'))['dados']
idx = {x['codigo']: x for x in d}
def norm(s): return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode().upper()
def find(p):
    for x in d:
        if re.search(p, norm(x['descricao'])): return x
    raise SystemExit('nao achou: ' + p)

PED = idx['88309']['custoUnitario']; SER = idx['88316']['custoUnitario']
alv = idx['103339']; cinta = idx['105033']; graute = idx['105792']
esc = idx['93358']; ent = idx['106122']; lastro = idx['100324']; cim = idx['98679']
calc = idx['94992']; dem = idx['104789']
for x in (alv, cinta, graute, esc, ent, lastro, cim, calc, dem):
    print(x['codigo'], x['unidade'], 'MO', x['custoMO'], x['descricao'][:90])

BDI = 0.28
def horas(c, cod):
    return sum(i['coeficiente'] for i in c['insumos'] if i['codigo'] == cod)
def linha(c, fator=1.0):
    return {'mo': round(c['custoMO'] * fator, 2), 'ped': horas(c, '88309') * fator, 'ser': horas(c, '88316') * fator}

# Itens: (n, descricao, unidade, qtd, estimado?, base, obs)
L = []
# (n, descricao, unidade, qtd, estimado?, base_horas, referencia, preco_unit_negociado ou None)
b = linha(alv); L.append(('1', 'Execução de alvenaria em bloco de concreto frisado (aparente) 14×19×39 cm, 4 fiadas sendo a última em canaleta, blocos fornecidos pela contratante: assentamento com argamassa, prumo, nível, amarração e juntas frisadas', 'm²', 35.0, False, b, 'Valor negociado (ref. SINAPI ' + alv['codigo'] + ')', 45.00))
GR_M3_POR_M = 0.015  # m³ de graute por metro de célula vertical (bloco 14 cm)
b1 = linha(cinta, 32.0 / 35.0); b2 = linha(graute, GR_M3_POR_M * 8.4 / 35.0)
b = {'mo': round(b1['mo'] + b2['mo'], 2), 'ped': b1['ped'] + b2['ped'], 'ser': b1['ser'] + b2['ser']}
L.append(('2', 'Enchimento de graute nas canaletas da última fiada (32 m) e nos pilares/células verticais a cada 3 m, com armação simples, limpeza das células e adensamento (graute e aço pela contratante)', 'm²', 35.0, False, b, 'Valor negociado (ref. SINAPI ' + cinta['codigo'] + ' + ' + graute['codigo'] + ')', 20.00))
b = linha(esc); L.append(('3', 'Escavação manual de vala (aprox. 30 m lineares, seção adotada 0,40 × 0,60 m = 7,20 m³), com material depositado ao lado', 'm³', 7.2, True, b, 'SINAPI ' + esc['codigo'], None))
b = {'mo': 0.0, 'ped': 0.0, 'ser': 8.0}
L.append(('4', 'Carga manual de entulho nas 2 caçambas estacionárias (5 m³ cada): 1 diária de equipe com 3 ajudantes', 'diária de ajudante', 3.0, False, b, 'Valor negociado (diária de ajudante)', 200.00))
b1 = linha(lastro, 0.10); b2 = linha(cim)
b = {'mo': round(b1['mo'] + b2['mo'], 2), 'ped': b1['ped'] + b2['ped'], 'ser': b1['ser'] + b2['ser']}
L.append(('5', 'Recondicionamento da base com areia e brita (lastro granular, e ≈ 10 cm) e recomposição do piso cimentado (traço 1:5, e = 2 cm) na área hoje cimentada', 'm²', 20.0, True, b, 'SINAPI ' + lastro['codigo'] + ' + ' + cim['codigo'], None))
b = linha(ent); L.append(('6', 'Retirada dos montes de terra e grama depositados sobre o gramado do campo, com carga manual e acondicionamento em caçamba/local indicado', 'm³', 6.0, True, b, 'SINAPI ' + ent['codigo'] + ' (por analogia)', None))
b1 = linha(dem, 0.06); b2 = linha(calc)
b = {'mo': round(b1['mo'] + b2['mo'], 2), 'ped': b1['ped'] + b2['ped'], 'ser': b1['ser'] + b2['ser']}
L.append(('7', 'Recomposição de calçada em concreto moldado in loco (≈ 10 m², e = 6 cm, acabamento convencional), incluindo demolição manual do trecho danificado', 'm²', 10.0, False, b, 'SINAPI ' + dem['codigo'] + ' + ' + calc['codigo'], None))

VALOR_GLOBAL = 4000.00  # valor fechado com o cliente; a diferença vira desconto comercial abatido do BDI

def moeda(v): return 'R$ ' + ('{:,.2f}'.format(v)).replace(',', 'X').replace('.', ',').replace('X', '.')
def num(v, c=2): return ('{:,.%df}' % c).format(v).replace(',', 'X').replace('.', ',').replace('X', '.')
def esc_(s): return html.escape(s)

rows = []; subtotal = 0; hp = 0; hs = 0
for n, desc, un, q, est, b, base, pneg in L:
    pu = pneg if pneg is not None else round(b['mo'] * (1 + BDI), 2)
    tot = round(pu * q, 2); subtotal += tot
    hp += b['ped'] * q; hs += b['ser'] * q
    rows.append((n, desc, un, q, est, pu, tot, base, b))
diaria = round((PED + SER) * 8 * (1 + BDI), 2)
dias = math.ceil(max(hp, hs) / 8 * 1.15)  # 1 pedreiro + 1 ajudante, 15% folga
dias2 = math.ceil(max(hp / 8, hs / 16) * 1.15)  # 1 pedreiro + 2 ajudantes
desconto = round(subtotal - VALOR_GLOBAL, 2); total = VALOR_GLOBAL
print('subtotal', subtotal, 'desconto', desconto, 'total', total, 'hp', hp, 'hs', hs, 'dias', dias, 'diaria', diaria)

hoje = datetime.date(2026, 9, 2); validade = hoje + datetime.timedelta(days=15)
fmt = lambda dt: dt.strftime('%d/%m/%Y')
NUM = 'PC-2026-0902-01'
EMP = 'RA Engenharia Especial LTDA'; CNPJ = '59.507.116/0001-64'
END = 'Rua Ovídio Bradamante de Toledo, 100, Apto 101 Bl. B — Tubalina — Uberlândia/MG'
RT = 'Eng. Rogério Alves de Souza — CREA-MG 323736'; ZAP = '(34) 9286-9383'; MAIL = 'contato@raengenhariaespecial.com.br'

def tab_rows():
    out = []
    for n, desc, un, q, est, pu, tot, base, b in rows:
        flag = ' <span class="est" title="quantidade estimada — a confirmar em visita técnica">◆</span>' if est else ''
        out.append('<tr><td class="c">%s</td><td>%s<div class="base">%s</div></td><td class="c">%s</td><td class="r">%s%s</td><td class="r">%s</td><td class="r">%s</td></tr>' % (n, esc_(desc), esc_(base), un, num(q), flag, moeda(pu), moeda(tot)))
    return ''.join(out)

def mem_rows():
    out = []
    for n, desc, un, q, est, pu, tot, base, b in rows:
        out.append('<tr><td class="c">%s</td><td>%s</td><td class="c">%s</td><td class="r">%s</td><td class="r">%s</td><td class="r">%s</td><td class="r">%s</td></tr>' % (n, esc_(base), un, num(b['ped'], 3), num(b['ser'], 3), (moeda(b['mo']) if b['mo'] else '—'), moeda(pu)))
    return ''.join(out)

WM = '<div class="wm">RA ENGENHARIA</div>'
ROD = '<div class="pg-rod"><span>%s · CNPJ %s · %s</span><span class="np"></span></div>' % (EMP, CNPJ, ZAP)
def pg(tit, corpo): return '<section class="pg interna">' + WM + '<div class="pg-head"><span class="pg-emp">RA ENGENHARIA</span><span class="pg-num">Proposta ' + NUM + '</span></div><h2 class="pg-tit">' + tit + '</h2>' + corpo + ROD + '</section>'

P = []
P.append('''<section class="pg capa">
  <div class="capa-top"><div class="logo"><span class="ra">RA</span><span class="eng">ENGENHARIA</span><small>ESPECIAL</small></div></div>
  <div class="capa-mid"><div class="kicker">PROPOSTA COMERCIAL</div>
    <h1>Prestação de mão de obra especializada de pedreiro e ajudante</h1>
    <div class="capa-obra">Alvenaria em bloco de concreto frisado com canaleta e graute, escavação manual, remoção de entulho, recomposição de base cimentada e de calçada</div></div>
  <div class="capa-info">
    <div class="ci-row"><span>Cliente</span><b>MY Engenharia</b></div>
    <div class="ci-row"><span>Local da obra</span><b>[ENDEREÇO DA OBRA — a informar]</b></div>
    <div class="ci-row"><span>Proposta nº</span><b>%s</b></div>
    <div class="ci-row"><span>Data</span><b>%s</b></div>
    <div class="ci-row"><span>Validade</span><b>15 dias (até %s)</b></div>
    <div class="ci-row"><span>Valor global (mão de obra)</span><b class="verde">%s</b></div>
  </div>
  <div class="capa-rod">%s · CNPJ %s<br>%s<br>%s · WhatsApp %s · %s</div>
</section>''' % (NUM, fmt(hoje), fmt(validade), moeda(total), EMP, CNPJ, END, RT, ZAP, MAIL))

P.append(pg('1. Apresentação', '''
<p>A <b>RA Engenharia Especial</b> atua na execução e gestão de obras civis, com responsabilidade técnica de engenheiro registrado no CREA-MG, equipe própria e foco em qualidade de execução, segurança e cumprimento de prazo.</p>
<p>Apresentamos à <b>MY Engenharia</b> nossa proposta para <b>prestação de mão de obra especializada de pedreiro e ajudante</b>, em regime de empreitada por preço unitário, para os serviços relacionados a seguir. Todo o material de consumo e aplicação (blocos, cimento, areia, brita, concreto, caçambas) é de fornecimento da contratante, conforme a solicitação recebida.</p>
<h3>2. Entendimento do escopo</h3>
<p>Os serviços solicitados foram interpretados da seguinte forma:</p>
<ul>
<li><b>Alvenaria — 35 m² em bloco de concreto frisado 14×19×39 cm:</b> 4 fiadas, sendo a última em bloco canaleta, ao longo de 32 m. Inclui assentamento com argamassa, prumo, nível, amarração, juntas frisadas, <b>grauteamento vertical a cada 3 m</b> e <b>concretagem da última fiada (canaleta)</b>. Blocos, canaletas, argamassa, graute, concreto e armação são fornecidos pela contratante; a frente inicia quando o material estiver completo na obra.</li>
<li><b>Escavação manual — aprox. 30 m lineares:</b> abertura de vala manual com o material escavado depositado ao lado. Como a seção não foi informada, adotamos 0,40 × 0,60 m (7,20 m³) como referência; a medição final será pelo volume efetivamente escavado.</li>
<li><b>Entulho — 2 caçambas:</b> carga manual do entulho existente em duas caçambas estacionárias de 5 m³ (10 m³), fornecidas pela contratante.</li>
<li><b>Recondicionamento de areia e brita na parte cimentada:</b> regularização da base com lastro de areia e brita e recomposição do piso cimentado na área indicada.</li>
<li><b>Retirada de montes de terra e grama sobre o campo:</b> remoção manual dos montes depositados sobre o gramado, com carga em caçamba ou transporte até local indicado no próprio terreno.</li>
<li><b>Recomposição de calçada — aprox. 10 m²:</b> demolição do trecho danificado e execução de novo passeio em concreto moldado no local.</li>
</ul>
<p class="nota">Os itens marcados com <span class="est">◆</span> na planilha têm quantidade <b>estimada</b>, a confirmar em visita técnica. Os preços unitários são firmes; o valor final será apurado por medição.</p>'''))

P.append(pg('3. Planilha de serviços — mão de obra', '''
<table class="prop-tbl"><thead><tr><th class="c">Item</th><th>Serviço</th><th class="c">Un.</th><th class="r">Qtd.</th><th class="r">Preço unit.</th><th class="r">Total</th></tr></thead>
<tbody>%s</tbody>
<tfoot><tr class="sub"><td colspan="5">Subtotal — mão de obra</td><td class="r">%s</td></tr>
<tr class="desc"><td colspan="5">Desconto comercial — abatimento do BDI (%s%%)</td><td class="r">− %s</td></tr>
<tr><td colspan="5">VALOR GLOBAL — MÃO DE OBRA (materiais e caçambas pela contratante)</td><td class="r">%s</td></tr></tfoot></table>
<p class="nota"><span class="est">◆</span> Quantidade estimada, a confirmar em visita técnica. Os itens 1, 2 e 4 têm preço negociado; os demais são referenciados no SINAPI-MG 06/2026 com BDI de %s%%, sobre o qual incide o desconto comercial do fechamento. Preços incluem encargos sociais e complementares (EPI, alimentação e transporte da equipe) e ferramentas manuais.</p>
''' % (tab_rows(), moeda(subtotal), num(desconto / subtotal * 100, 1), moeda(desconto), moeda(total), int(BDI * 100))))

P.append(pg('4. Está incluso / Não está incluso', '''
<div class="cols"><div><h3>✔ Incluso</h3><ul>
<li>Mão de obra de pedreiro e ajudante, com encargos sociais e complementares.</li>
<li>Ferramentas manuais e equipamentos leves da equipe (colher, prumo, nível, carrinho, pá, enxada, picareta, régua, desempenadeira).</li>
<li>EPIs da equipe (capacete, luvas, botas, óculos), alimentação e transporte.</li>
<li>Preparo de argamassa e concreto em obra (manual ou em betoneira fornecida pela contratante).</li>
<li>Acompanhamento técnico de engenheiro responsável (CREA-MG) e medição dos serviços.</li>
<li>Limpeza da área de trabalho ao final de cada frente.</li>
</ul></div><div><h3>✖ Não incluso</h3><ul>
<li>Materiais: blocos, cimento, cal, areia, brita, concreto, aço, madeira e demais insumos.</li>
<li>Locação, entrega e retirada das caçambas de entulho e destinação final (CTR).</li>
<li>Água e energia elétrica no local.</li>
<li>Equipamentos pesados (retroescavadeira, compactador, martelete), betoneira e andaimes.</li>
<li>Projetos, ART de projeto, taxas e licenças municipais.</li>
<li>Serviços não relacionados na planilha (chapisco, reboco, pintura, instalações), que podem ser orçados à parte ou executados por diária.</li>
</ul></div></div>
<h3>5. Premissas</h3>
<ul>
<li>Acesso livre à obra em horário comercial (segunda a sexta, 7h às 17h) e local para guarda de ferramentas.</li>
<li>Materiais entregues na obra antes do início de cada frente; paralisação por falta de material superior a 2 dias poderá gerar remobilização ou cobrança por diária, mediante acordo prévio.</li>
<li>Quantidades marcadas com ◆ e a seção da vala serão confirmadas em visita técnica antes do início, sem alteração dos preços unitários.</li>
<li>Escavação em solo comum, sem rocha, sem lençol freático e sem interferências (tubulações, cabos) não sinalizadas pela contratante.</li>
<li>Serviços em conformidade com as normas ABNT aplicáveis (NBR 8545 — alvenaria, NBR 9061 — escavações, NBR 9050 — calçadas) e NR-18.</li>
</ul><h3>Opção complementar — equipe por diária</h3>
<p>Para serviços extras ou não previstos na planilha, oferecemos a equipe (1 pedreiro + 1 ajudante, 8 h/dia) ao valor de <b>%s por dia</b>, medido por dia efetivamente trabalhado e autorizado por escrito pela contratante.</p>
''' % moeda(diaria)))

P.append(pg('6. Metodologia e cronograma', '''
<p>Execução por frentes de serviço, na ordem que melhor aproveita a equipe e as caçambas disponíveis:</p>
<ol>
<li><b>Mobilização e visita técnica</b> — conferência das quantidades ◆, marcação da vala e da calçada, definição do local de descarte dos montes de terra.</li>
<li><b>Limpezas e cargas</b> — retirada dos montes de terra e grama do campo e carga do entulho nas 2 caçambas (aproveitando as mesmas caçambas para o entulho da calçada).</li>
<li><b>Escavação manual</b> — abertura da vala nos 30 m, com material depositado ao lado para reaterro pela contratante ou reaproveitamento.</li>
<li><b>Base cimentada</b> — recondicionamento da base com areia e brita, apiloamento e recomposição do piso cimentado.</li>
<li><b>Calçada</b> — demolição do trecho danificado, regularização, formas simples e concretagem com acabamento desempenado.</li>
<li><b>Alvenaria</b> — assim que os blocos estiverem completos: marcação e 1ª fiada, elevação das fiadas com juntas frisadas, grauteamento vertical a cada 3 m, assentamento da canaleta, armação e concretagem da última fiada (32 m).</li>
<li><b>Medição final e entrega</b> — conferência conjunta com a contratante e limpeza da área.</li>
</ol>
<h3>Prazo estimado</h3>
<p>Com uma equipe de 1 pedreiro e 1 ajudante, estimamos <b>%d dias úteis</b> de execução para o conjunto dos serviços; com 1 pedreiro e 2 ajudantes o prazo cai para cerca de <b>%d dias úteis</b>. Ambos condicionados à disponibilidade dos materiais e das caçambas na obra. O cronograma detalhado será alinhado na reunião de início.</p>
<h3>7. Condições comerciais</h3>
<table class="prop-tbl"><tbody>
<tr><td><b>Regime</b></td><td>Empreitada por valor global de <b>R$ 4.000,00</b> para o escopo da planilha. Quantidades marcadas com ◆ serão conferidas em visita técnica; variação superior a 20%% em qualquer item será tratada por aditivo ou por diária.</td></tr>
<tr><td><b>Pagamento</b></td><td>50%% (R$ 2.000,00) na mobilização e 50%% (R$ 2.000,00) na conclusão, contra vistoria final aprovada. Serviços por diária: fechamento semanal.</td></tr>
<tr><td><b>Forma</b></td><td>PIX ou transferência bancária, com emissão de nota fiscal de serviço.</td></tr>
<tr><td><b>Reajuste</b></td><td>Preços fixos pelo prazo de validade; após 15 dias, sujeitos a revisão pela tabela SINAPI-MG vigente.</td></tr>
<tr><td><b>Validade</b></td><td>15 dias a contar de %s.</td></tr>
<tr><td><b>Início</b></td><td>Até 5 dias úteis após o aceite e a confirmação da entrega dos materiais.</td></tr>
</tbody></table>''' % (dias, dias2, fmt(hoje))))

P.append(pg('8. Responsabilidades, garantias e aceite', '''
<div class="cols"><div><h3>Da RA Engenharia</h3><ul>
<li>Executar os serviços com qualidade técnica, dentro das normas e do prazo acordado.</li>
<li>Manter a equipe com EPIs e cumprir a NR-18.</li>
<li>Designar engenheiro responsável para acompanhamento e medição.</li>
<li>Comunicar imediatamente qualquer interferência ou impedimento encontrado.</li>
</ul></div><div><h3>Da MY Engenharia</h3><ul>
<li>Fornecer materiais, caçambas, água e energia nos prazos combinados.</li>
<li>Informar o local exato dos serviços e eventuais interferências enterradas.</li>
<li>Aprovar as medições e efetuar os pagamentos nas datas previstas.</li>
<li>Indicar um responsável para as definições em campo.</li>
</ul></div></div>
<h3>9. Garantias</h3>
<p>Garantia de execução de <b>90 dias</b> para os serviços de mão de obra (prumo, nível, aderência do cimentado e da calçada), contados da medição final, cobrindo defeitos de execução. Não cobre falhas decorrentes de material fornecido pela contratante, uso inadequado ou intervenção de terceiros.</p>
<h3>10. Aceite</h3>
<p>Estando de acordo, solicitamos a devolução desta proposta assinada, ou a confirmação por e-mail/WhatsApp, para agendamento da visita técnica e início dos serviços.</p>
<div class="assinaturas">
<div><div class="linha-assin"></div><div class="assin">%s<br><span>%s</span><br><span>%s · %s</span></div></div>
<div><div class="linha-assin"></div><div class="assin">MY Engenharia<br><span>[NOME DO RESPONSÁVEL]</span><br><span>Data: ____/____/______</span></div></div>
</div>''' % (EMP, RT, ZAP, MAIL)))

P.append(pg('Anexo — Memória de cálculo dos preços unitários', '''
<p class="nota">Referência: <b>SINAPI-MG, competência 06/2026</b>, regime desonerado, parcela de mão de obra das composições indicadas (materiais e equipamentos excluídos). Salários-hora com encargos complementares: pedreiro (88309) <b>%s/h</b> · servente (88316) <b>%s/h</b>. Preço unitário = custo MO × (1 + BDI %s%%).</p>
<table class="prop-tbl"><thead><tr><th class="c">Item</th><th>Composição de referência</th><th class="c">Un.</th><th class="r">h pedreiro</th><th class="r">h ajudante</th><th class="r">Custo MO</th><th class="r">Preço unit.</th></tr></thead>
<tbody>%s</tbody></table>
<ul class="nota">
<li>Itens 1 e 2: <b>preço negociado</b> — R$ 45,00/m² para a execução da alvenaria e R$ 20,00/m² para o enchimento de graute nas canaletas (32 m) e nos pilares a cada 3 m. Referências SINAPI %s (alvenaria aparente 14×19×39), 105033 (cinta em canaleta) e 105792 (graute vertical) usadas apenas para estimar as horas.</li>
<li>Item 3: <b>93358</b> — Escavação manual de vala (m³); seção adotada 0,40 × 0,60 m.</li>
<li>Item 4: <b>preço negociado</b> — 1 diária de 3 ajudantes a R$ 200,00 cada para encher as 2 caçambas.</li>
<li>Item 5: <b>100324</b> — lastro de brita (m³, convertido para m² com e = 10 cm) + <b>98679</b> — piso cimentado 1:5, e = 2 cm, liso.</li>
<li>Item 6: <b>106122</b> — remoção de entulho classe A com acondicionamento em caçamba, adotada <b>por analogia</b> para a carga manual (3,42 h de servente por m³).</li>
<li>Item 7: <b>104789</b> — demolição manual de piso de concreto (m³, convertido para m² com e = 6 cm) + <b>94992</b> — passeio em concreto moldado in loco, e = 6 cm.</li>
<li>Fechamento: subtotal %s − desconto comercial de %s (abatimento do BDI) = <b>valor global %s</b>.</li>
<li>Diária da equipe: (%s + %s) × 8 h × 1,%d = <b>%s</b>.</li>
<li>Horas totais estimadas: pedreiro %s h · ajudante %s h → %d dias úteis com folga de 15%%.</li>
</ul>''' % (moeda(PED), moeda(SER), int(BDI * 100), mem_rows(), alv['codigo'], moeda(subtotal), moeda(desconto), moeda(total), moeda(PED), moeda(SER), int(BDI * 100), moeda(diaria), num(hp, 1), num(hs, 1), dias)))

CSS = '''
:root{--p-navy:#0f2740;--p-aco:#2e6f9e;--p-verde:#15803d;--p-linha:#d8e0ea;--raio:10px}
*{box-sizing:border-box}
body{margin:0;background:#2b3a4a;font-family:"Segoe UI","Helvetica Neue",Arial,sans-serif;color:#1a2632;-webkit-print-color-adjust:exact;print-color-adjust:exact}
.toolbar{position:fixed;top:0;left:0;right:0;height:52px;z-index:10;background:var(--p-navy);color:#fff;display:flex;align-items:center;gap:12px;padding:0 18px;font-size:14px}
.toolbar .ttl{flex:1;font-weight:700}
.toolbar button{background:#16a34a;color:#fff;border:0;border-radius:8px;padding:8px 16px;font-weight:700;cursor:pointer}
.doc{padding:70px 0 40px}
.pg{width:210mm;min-height:297mm;background:#fff;margin:0 auto 18px;box-shadow:0 6px 24px rgba(0,0,0,.3);padding:22mm 20mm 24mm;position:relative;font-size:11pt;line-height:1.5;overflow:hidden}
.pg .wm{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;font-size:64pt;font-weight:800;color:rgba(15,39,64,.045);transform:rotate(-30deg);pointer-events:none;letter-spacing:6px;z-index:0}
.pg>*:not(.wm){position:relative;z-index:1}
.pg-head{display:flex;justify-content:space-between;font-size:8.5pt;letter-spacing:2px;color:var(--p-aco);font-weight:700;margin-bottom:10px}
.pg-head .pg-num{color:#7a8a99;letter-spacing:.3px;font-weight:600}
.pg h2.pg-tit{font-size:17pt;color:var(--p-navy);border-bottom:3px solid var(--p-aco);padding-bottom:8px;margin:0 0 14px;position:relative}
.pg h2.pg-tit::after{content:"";position:absolute;left:0;bottom:-3px;width:58px;height:3px;background:#16a34a}
.pg h3{font-size:12pt;color:var(--p-aco);margin:14px 0 6px}
.pg p{margin:6px 0}
.pg ul,.pg ol{margin:6px 0 6px 18px;padding:0}.pg li{margin:3px 0}
.pg .cols{display:grid;grid-template-columns:1fr 1fr;gap:24px}
.pg .nota{font-size:9.5pt;color:#6b7b8a}
.pg .pg-rod{position:absolute;bottom:12mm;left:20mm;right:20mm;border-top:1px solid var(--p-linha);padding-top:6px;font-size:8.5pt;color:#7a8a99;display:flex;justify-content:space-between}
.est{color:#b45309;font-weight:700}
.pg.capa{background:linear-gradient(155deg,#0b1d31 0%,#143454 58%,#1c4b73 128%);color:#fff;display:flex;flex-direction:column;padding:24mm 20mm}
.pg.capa::before{content:"";position:absolute;inset:0;z-index:0;background-image:linear-gradient(rgba(255,255,255,.05) 1px,transparent 1px),linear-gradient(90deg,rgba(255,255,255,.05) 1px,transparent 1px);background-size:42px 42px;-webkit-mask-image:radial-gradient(560px 460px at 82% 16%,#000,transparent 74%);mask-image:radial-gradient(560px 460px at 82% 16%,#000,transparent 74%)}
.pg.capa::after{content:"";position:absolute;right:-120px;top:-100px;width:420px;height:420px;border-radius:50%;background:radial-gradient(circle,rgba(111,208,138,.16),transparent 70%);z-index:0}
.pg.capa>*{position:relative;z-index:1}
.logo{display:inline-flex;align-items:baseline;gap:8px;border-left:5px solid #6fd08a;padding-left:14px;line-height:1}
.logo .ra{font-size:34pt;font-weight:800;letter-spacing:-1px}.logo .eng{font-size:16pt;font-weight:700;letter-spacing:4px;color:#cfe0ef}.logo small{font-size:8pt;letter-spacing:3px;color:#9fb6cc}
.capa-top{margin-bottom:auto}
.capa-mid .kicker{letter-spacing:6px;font-size:11pt;color:#9be7af;font-weight:700}
.capa-mid .kicker::after{content:"";display:block;width:58px;height:4px;background:#16a34a;border-radius:3px;margin-top:14px}
.capa-mid h1{color:#fff;font-size:26pt;font-weight:800;line-height:1.15;letter-spacing:-.5px;margin:18px 0 10px}
.capa-obra{font-size:12.5pt;color:#cfe0ef}
.capa-info{background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.13);border-radius:var(--raio);padding:16px 22px;margin:28px 0}
.ci-row{display:flex;justify-content:space-between;align-items:baseline;gap:20px;padding:7px 0;border-bottom:1px solid rgba(255,255,255,.1);font-size:11pt}
.ci-row:last-child{border-bottom:none;padding-bottom:0}.ci-row:last-child b{font-size:16pt}
.ci-row span{color:#a9c4dc;white-space:nowrap}.ci-row b{text-align:right;font-variant-numeric:tabular-nums}.verde{color:#6fd08a}
.capa-rod{margin-top:auto;font-size:9pt;color:rgba(255,255,255,.65);border-top:1px solid rgba(255,255,255,.1);padding-top:14px;line-height:1.6}
table.prop-tbl{width:100%;border-collapse:collapse;font-size:9.4pt;margin:6px 0;line-height:1.38}
table.prop-tbl th,table.prop-tbl td{padding:5px 7px;border-bottom:1px solid var(--p-linha);text-align:left;vertical-align:top}
table.prop-tbl th{background:var(--p-navy);color:#fff;font-weight:700;letter-spacing:.3px}
table.prop-tbl tbody tr:nth-child(even) td{background:#f7fafd}
table.prop-tbl td.r,table.prop-tbl th.r{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
table.prop-tbl td.c,table.prop-tbl th.c{text-align:center}
table.prop-tbl .base{font-size:8pt;color:#7a8a99;margin-top:2px}
table.prop-tbl tfoot td{background:#eef4fa;font-weight:700;border-top:2px solid var(--p-navy);color:var(--p-navy)}
table.prop-tbl tfoot tr.sub td{background:#fff;border-top:2px solid var(--p-navy);font-weight:600;color:#1a2632}
table.prop-tbl tfoot tr.desc td{background:#fff;border-top:0;font-weight:600;color:#b45309}
.assinaturas{display:grid;grid-template-columns:1fr 1fr;gap:40px;margin-top:56px;text-align:center}
.assin{font-weight:600;line-height:1.5}.assin span{color:#6b7b8a;font-weight:400;font-size:9.5pt}
.linha-assin{border-top:1.5px solid #1a2632;margin-bottom:8px}
@page{size:A4;margin:0}
@media print{body{background:#fff}.toolbar{display:none}.doc{padding:0}.pg{margin:0;box-shadow:none;page-break-after:always;min-height:297mm;height:297mm}.pg:last-child{page-break-after:auto}}
'''
HTML = '''<!DOCTYPE html><html lang="pt-BR"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Proposta %s — MY Engenharia — RA Engenharia</title><style>%s</style></head><body>
<div class="toolbar"><span class="ttl">Proposta Comercial %s · MY Engenharia · RA Engenharia</span><button onclick="window.print()">Imprimir / Salvar PDF</button></div>
<div class="doc">%s</div></body></html>''' % (NUM, CSS, NUM, ''.join(P))
open('propostas/2026-09-02-MY-Engenharia-mao-de-obra-PC-2026-0902-01.html', 'w').write(HTML)
print('ok html')

# =====================================================================
# PACOTE DE ORÇAMENTO — o mesmo orçamento pronto para subir no OrçaPRO
# (app/?importar=<url> ou 💾 Backup › Restaurar). Ver propostas/LEIA-ME.md.
#
# O app não tem "desconto" no fechamento, e zero/negativo não é preço. Então
# o pacote leva os PREÇOS FINAIS: cada unitário recebe o mesmo fator do
# desconto (4.000 / 5.642,64) e uma busca de centavos garante que a soma
# feche exatamente em R$ 4.000,00 com 2 casas. BDI 0 % (o preço já é venda).
# A memória de cada item registra o unitário antes do desconto.
# =====================================================================
import sys, os, itertools
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from orcapro_pacote import Pacote, bdi_zero

fator = VALOR_GLOBAL / subtotal
qs = [r[3] for r in rows]                       # quantidades
base_c = [int(round(r[5] * fator * 100)) for r in rows]   # unitários em centavos, arredondados
base_c = [int(5 * round(c / 5.0)) if q == 7.2 else c for c, q in zip(base_c, qs)]  # 7,2 m³ × múltiplo de 5 centavos = centavos inteiros
def soma_c(cs): return sum(int(round(c * q)) for c, q in zip(cs, qs))
alvo_c = int(round(VALOR_GLOBAL * 100)); melhor = None
# passos que mantêm o total em centavos exatos: escavação (q=7,2) só anda de 5 em 5 centavos
passos = [range(-4, 5) if q != 7.2 else range(-20, 25, 5) for q in qs]
for delta in itertools.product(*passos):
    cs = [b + d for b, d in zip(base_c, delta)]
    if soma_c(cs) != alvo_c: continue
    custo = sum(abs(d) for d in delta)
    if melhor is None or custo < melhor[0]: melhor = (custo, cs)
    if custo <= 2: break
assert melhor, 'não achei combinação de centavos que feche em ' + str(VALOR_GLOBAL)
pu_final = [c / 100.0 for c in melhor[1]]

P = Pacote(gerado_por='propostas/gerar-proposta-my-engenharia.py — proposta ' + NUM)
cli = P.cliente('MY Engenharia', uf='MG', origem='indicacao', obs='Cadastro criado pela proposta ' + NUM + '. Complete CNPJ, telefone e e-mail.')
obra = P.obra('MY Engenharia — mão de obra (alvenaria, escavação, entulho e calçada)', cliente=cli,
              local='[ENDEREÇO DA OBRA — a informar]', tipo='reforma', status='planejamento', valor=VALOR_GLOBAL,
              obs='Escopo: 35 m² alvenaria bloco 14x19x39 frisado com canaleta e graute, ~30 m escavação manual, 2 caçambas de entulho, base cimentada, retirada de terra/grama, calçada 10 m².')
orc = P.orcamento(NUM, 'Mão de obra pedreiro e ajudante — MY Engenharia', cliente=cli, obra=obra, uf='MG',
                  competencia='2026-06', bdi=bdi_zero(), categoria='Mão de obra', prazo_entrega=str(dias) + ' dias úteis',
                  comercial={
                      'apresentacao': 'Prestação de mão de obra especializada de pedreiro e ajudante, em regime de empreitada por valor global de ' + moeda(VALOR_GLOBAL) + '. Materiais, caçambas, água e energia por conta da contratante. Preços unitários já com o desconto comercial de fechamento (subtotal ' + moeda(subtotal) + ' − ' + moeda(desconto) + ').',
                      'condicoesPagamento': '50% (R$ 2.000,00) na mobilização e 50% (R$ 2.000,00) na conclusão, contra vistoria final aprovada. Serviços extras por diária (1 pedreiro + 1 ajudante): ' + moeda(diaria) + '/dia, fechamento semanal. PIX ou transferência, com NF de serviço.',
                      'prazoExecucao': str(dias) + ' dias úteis com 1 pedreiro + 1 ajudante (cerca de ' + str(dias2) + ' com 2 ajudantes), condicionado à entrega dos materiais.',
                      'validadeProposta': '15 dias corridos a contar de ' + fmt(hoje) + '.',
                      'garantia': 'Garantia de execução de 90 dias para os serviços de mão de obra, contados da vistoria final; não cobre falhas de material fornecido pela contratante.',
                      'incluso': 'Mão de obra de pedreiro e ajudante com encargos sociais e complementares;\nFerramentas manuais e equipamentos leves;\nEPIs, alimentação e transporte da equipe;\nPreparo de argamassa e concreto em obra;\nAcompanhamento técnico (CREA-MG) e medição;\nLimpeza da área ao final de cada frente.',
                      'excluso': 'Materiais (blocos, cimento, areia, brita, concreto, aço, graute);\nLocação, retirada e destinação das caçambas;\nÁgua e energia no local;\nEquipamentos pesados, betoneira e andaimes;\nProjetos, ART de projeto, taxas e licenças;\nServiços não relacionados na planilha.'
                  })
grupos = [('Alvenaria e graute', ['1', '2']), ('Escavação e limpeza', ['3', '4', '6']), ('Pisos e calçada', ['5', '7'])]
por_n = {r[0]: (r, pu) for r, pu in zip(rows, pu_final)}
for nome_et, ns in grupos:
    et = P.etapa(orc, nome_et)
    for n in ns:
        r, pu = por_n[n]
        desc = r[1].split(' (SINAPI')[0]
        memoria = ('Referência: ' + r[7] + '. Unitário da proposta antes do desconto: ' + moeda(r[5]) +
                   ' → com desconto comercial de ' + num(desconto / subtotal * 100, 1) + '%: ' + moeda(pu) + '.' +
                   (' Quantidade ESTIMADA, a confirmar em visita técnica.' if r[4] else ''))
        P.item(orc, et, 'MO-0' + n, desc, r[2], r[3], pu, memoria=memoria)
assert abs(P.total(orc) - VALOR_GLOBAL) < 0.005, P.total(orc)
# ---- Planilha Excel da proposta (o botão "Abrir a planilha" do PDF aponta para ela) ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
XLSX = 'propostas/2026-09-02-MY-Engenharia-mao-de-obra-PC-2026-0902-01.xlsx'
URL_XLSX = 'https://ra-engenharia.github.io/orcapro/' + XLSX
wb = Workbook(); ws = wb.active; ws.title = 'Proposta'
navy = PatternFill('solid', fgColor='0F3B5E'); branco = Font(bold=True, color='FFFFFF'); neg = Font(bold=True)
fino = Side(style='thin', color='CBD5E1'); borda = Border(top=fino, left=fino, right=fino, bottom=fino)
ws['A1'] = 'RA Engenharia Especial LTDA — Proposta ' + NUM; ws['A1'].font = Font(bold=True, size=14, color='0F3B5E')
ws['A2'] = 'Cliente: MY Engenharia'; ws['A3'] = 'Objeto: prestação de mão de obra de pedreiro e ajudante'
ws['A4'] = 'Data: ' + fmt(hoje) + ' · validade 15 dias · referência SINAPI-MG 06/2026'
cab = ['Item', 'Serviço', 'Un.', 'Qtd.', 'Preço unit. (R$)', 'Total (R$)', 'Referência']
ws.append([]); ws.append(cab)
for c in ws[6]: c.fill = navy; c.font = branco; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = borda
lin = 7
for n, desc, un, q, est, pu, tot, base, b in rows:
    ws.append([n, desc + (' (quantidade estimada)' if est else ''), un, q, pu, None, base])
    ws.cell(row=lin, column=6).value = '=D%d*E%d' % (lin, lin)
    for col in range(1, 8): ws.cell(row=lin, column=col).border = borda
    ws.cell(row=lin, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    for col in (5, 6): ws.cell(row=lin, column=col).number_format = '#,##0.00'
    lin += 1
fim = lin - 1
ws.append(['', 'Subtotal', '', '', '', '=SUM(F7:F%d)' % fim]); ws.cell(row=lin, column=6).number_format = '#,##0.00'; ws.cell(row=lin, column=2).font = neg; lin += 1
ws.append(['', 'Desconto comercial (abatimento do BDI)', '', '', '', -desconto]); ws.cell(row=lin, column=6).number_format = '#,##0.00'; lin += 1
ws.append(['', 'VALOR GLOBAL', '', '', '', '=F%d+F%d' % (lin - 2, lin - 1)]); ws.cell(row=lin, column=6).number_format = '#,##0.00'
for col in range(1, 8): ws.cell(row=lin, column=col).font = neg; ws.cell(row=lin, column=col).fill = PatternFill('solid', fgColor='EEF4FA')
for col, w in zip('ABCDEFG', (6, 70, 8, 9, 16, 16, 34)): ws.column_dimensions[col].width = w
ws.freeze_panes = 'A7'
m = wb.create_sheet('Memória de cálculo')
m.append(['Item', 'Composição de referência', 'Un.', 'h pedreiro', 'h ajudante', 'Custo MO SINAPI (R$)', 'Preço unit. proposta (R$)'])
for c in m[1]: c.fill = navy; c.font = branco
for n, desc, un, q, est, pu, tot, base, b in rows:
    m.append([n, base, un, round(b['ped'], 3), round(b['ser'], 3), b['mo'] or None, pu])
m.append([]); m.append(['Salário-hora com encargos: pedreiro (88309) R$ %.2f · servente (88316) R$ %.2f · BDI %d%% · SINAPI-MG 06/2026 desonerado' % (PED, SER, int(BDI * 100))])
for col, w in zip('ABCDEFG', (6, 46, 8, 12, 12, 20, 22)): m.column_dimensions[col].width = w
wb.save(XLSX); print('ok xlsx', XLSX)
orc['comercial']['linkPlanilha'] = URL_XLSX

arq = P.salvar('propostas/2026-09-02-MY-Engenharia-mao-de-obra-PC-2026-0902-01.orcapro.json')
print('ok pacote', arq, 'total', P.total(orc), 'unitarios', pu_final)
